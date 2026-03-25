import streamlit as st
import pandas as pd
import numpy as np
import datetime
import io
import bcrypt
import json
import os
from pathlib import Path
from sqlalchemy import create_engine, text
from usuarios_config import USUARIOS_CREDENCIALES, CREDENCIALES_INICIALES

# Version: 4.0 - SQLite persistent backend

# ----------------------------
# CONFIG
# ----------------------------
st.set_page_config(page_title="Inventarios Rotativos - Grupo Cenoa", layout="wide", page_icon="📦")

# DB configurable: PostgreSQL/Supabase por DATABASE_URL, SQLite como fallback local
DB_PATH = Path(__file__).resolve().parent / "inventarios.db"
DATABASE_URL = (
    os.getenv("DATABASE_URL")
    or st.secrets.get("database", {}).get("url")
    or f"sqlite:///{DB_PATH.as_posix()}"
)
DB_BACKEND = "SQLite" if DATABASE_URL.startswith("sqlite") else "Externa"

SHEET_HIST = "Historial_Inventarios"
SHEET_DET = "Detalle_Articulos"
SHEET_AUDIT = "Audit_Log"
SHEET_BASE = "Base_Excel_Articulos"

# Columnas esperadas del Excel
C_ART = "Artículo"
C_LOC = "Locación"
C_DESC = "Descripción"
C_STOCK = "Stock"
C_COSTO = "Cto.Rep."

# Concesionarias y sucursales
CONCESIONARIAS = {
    "Autolux": ["Ax Jujuy", "Ax Salta", "Ax Tartagal", "Ax Lajitas", "Ax Taller Movil"],
    "Autosol": ["As Jujuy", "As Salta", "As Tartagal", "As Taller Express", "As Taller Movil"],
    "Ciel": ["Ac Jujuy"],
    "Portico": ["Las Lomas", "Brown"],
}

MODULE_META = {
    "nuevo": {
        "label": "1) Nuevo inventario",
        "icon": "📦",
        "title": "Nuevo Inventario",
        "description": "Importá el stock, generá la muestra ABC y creá un nuevo operativo de auditoría.",
    },
    "conteo": {
        "label": "2) Conteo físico",
        "icon": "🧮",
        "title": "Conteo Físico",
        "description": "Registrá el conteo del auditor y descargá la evidencia operativa en Excel.",
    },
    "justificaciones": {
        "label": "3) Justificaciones",
        "icon": "📝",
        "title": "Justificaciones y Validación",
        "description": "Gestioná diferencias, validaciones y decisiones de ajuste con trazabilidad.",
    },
    "cierre": {
        "label": "4) Cierre + Reporte",
        "icon": "✅",
        "title": "Cierre y Reporte",
        "description": "Consolidá resultados, evaluá exactitud y emití el reporte final del inventario.",
    },
    "dashboards": {
        "label": "5) Dashboards",
        "icon": "📊",
        "title": "Dashboards Ejecutivos",
        "description": "Visualizá KPIs clave, ranking de sucursales y desempeño general del proceso.",
    },
}

def inject_modern_theme():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;600;700&family=Inter:wght@400;500;600&display=swap');

        /* ─── BASE ─────────────────────────────────────────── */
        html, body, [class*="css"] {
            font-family: 'Inter', sans-serif;
        }

        .stApp {
            background: #f1f5f9;
            color-scheme: light;
            --primary-color: #b91c1c;
            --background-color: #f1f5f9;
            --secondary-background-color: #ffffff;
            --text-color: #0f172a;
            --body-text-color: #1e293b;
            --dataframe-header-background-color: #f8fafc;
            --dataframe-border-color: #e2e8f0;
            --dataframe-row-hover-color: #f1f5f9;
        }

        [data-testid="stAppViewContainer"] > .main {
            background: transparent;
        }

        .block-container {
            padding-top: 1.5rem;
            padding-bottom: 2.5rem;
            max-width: 1400px;
        }

        /* ─── TYPOGRAPHY ───────────────────────────────────── */
        h1, h2, h3, h4, h5, h6 {
            font-family: 'Space Grotesk', sans-serif !important;
            color: #0f172a !important;
            letter-spacing: -0.02em;
            font-weight: 700;
        }

        p, li, span, div, label {
            color: #1e293b;
        }

        .stMarkdown p {
            color: #334155;
        }

        /* ─── SIDEBAR ──────────────────────────────────────── */
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #0f172a 0%, #1e293b 100%);
            border-right: 1px solid rgba(255, 255, 255, 0.06);
        }

        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] span,
        [data-testid="stSidebar"] div,
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] small,
        [data-testid="stSidebar"] a {
            color: #cbd5e1 !important;
        }

        [data-testid="stSidebar"] h1,
        [data-testid="stSidebar"] h2,
        [data-testid="stSidebar"] h3 {
            color: #f1f5f9 !important;
        }

        /* ─── BRAND BLOCK ──────────────────────────────────── */
        .brand-shell {
            padding: 0.9rem 1rem 1rem 1rem;
            border-radius: 18px;
            background: linear-gradient(135deg, rgba(185, 28, 28, 0.32) 0%, rgba(13, 148, 136, 0.18) 100%);
            border: 1px solid rgba(255, 255, 255, 0.08);
            margin-bottom: 1rem;
        }

        .brand-kicker {
            font-size: 0.68rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.14em;
            color: #94a3b8 !important;
        }

        .brand-title {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 1.1rem;
            font-weight: 700;
            color: #f1f5f9 !important;
            line-height: 1.25;
            margin-top: 0.3rem;
        }

        /* ─── SIDEBAR BUTTONS ──────────────────────────────── */
        [data-testid="stSidebar"] div[data-testid="stButton"] > button {
            width: 100%;
            text-align: left !important;
            background: rgba(255, 255, 255, 0.05) !important;
            border: 1px solid rgba(255, 255, 255, 0.08) !important;
            border-radius: 12px !important;
            color: #e2e8f0 !important;
            font-weight: 500;
            font-size: 0.9rem;
            padding: 0.6rem 0.9rem !important;
            box-shadow: none !important;
            transition: all 0.15s ease;
            margin-bottom: 2px;
        }

        [data-testid="stSidebar"] div[data-testid="stButton"] > button:hover {
            background: rgba(185, 28, 28, 0.22) !important;
            border-color: rgba(185, 28, 28, 0.45) !important;
            color: #ffffff !important;
            transform: none;
        }

        /* ─── HERO CARD ────────────────────────────────────── */
        .hero-card {
            background: #ffffff;
            border: 1px solid #e2e8f0;
            border-radius: 20px;
            padding: 1.4rem 1.75rem;
            box-shadow: 0 2px 16px rgba(15, 23, 42, 0.07);
            margin-bottom: 1.5rem;
        }

        .hero-kicker {
            font-size: 0.7rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.13em;
            color: #b91c1c !important;
            margin-bottom: 0.5rem;
        }

        .hero-title {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 2rem;
            font-weight: 700;
            color: #0f172a !important;
            margin: 0;
            line-height: 1.1;
        }

        .hero-desc {
            margin-top: 0.6rem;
            color: #475569 !important;
            font-size: 1rem;
        }

        .hero-meta {
            display: flex;
            gap: 0.5rem;
            flex-wrap: wrap;
            margin-top: 0.9rem;
        }

        .hero-pill {
            padding: 0.3rem 0.75rem;
            border-radius: 999px;
            background: #f1f5f9;
            border: 1px solid #e2e8f0;
            font-size: 0.8rem;
            color: #334155 !important;
            font-weight: 500;
        }

        /* ─── MAIN AREA BUTTONS ────────────────────────────── */
        div[data-testid="stButton"] > button,
        div[data-testid="stDownloadButton"] > button {
            border-radius: 10px;
            border: 1.5px solid #e2e8f0;
            background: #ffffff;
            color: #0f172a !important;
            font-weight: 600;
            font-size: 0.88rem;
            padding: 0.5rem 1.1rem;
            transition: all 0.15s ease;
            box-shadow: 0 1px 4px rgba(15, 23, 42, 0.06);
        }

        div[data-testid="stButton"] > button:hover,
        div[data-testid="stDownloadButton"] > button:hover {
            border-color: #b91c1c;
            color: #b91c1c !important;
            box-shadow: 0 4px 14px rgba(185, 28, 28, 0.14);
            transform: translateY(-1px);
        }

        /* ─── INPUTS (text, number) ───────────────────────── */
        [data-baseweb="select"] > div,
        div[data-testid="stTextInputRootElement"] > div,
        div[data-testid="stNumberInputContainer"] {
            border-radius: 10px !important;
            border: 1.5px solid #e2e8f0 !important;
            background: #ffffff !important;
        }

        input, textarea, select {
            color: #0f172a !important;
            background: #ffffff !important;
        }

        /* Selected value shown in selectbox trigger */
        [data-baseweb="select"] [data-baseweb="tag"],
        [data-baseweb="select"] span,
        [data-baseweb="select"] div {
            color: #0f172a !important;
            background-color: transparent;
        }

        /* Multiselect tags (pills inside input) */
        [data-baseweb="tag"] {
            background-color: #e2e8f0 !important;
            border: 1px solid #cbd5e1 !important;
            border-radius: 6px !important;
            color: #1e293b !important;
        }

        [data-baseweb="tag"] span,
        [data-baseweb="tag"] div {
            color: #1e293b !important;
        }

        /* ─── DROPDOWN / SELECT POPUP ─────────────────────── */
        /* The floating popup card */
        [data-baseweb="popover"],
        [data-baseweb="popover"] > div,
        [data-baseweb="menu"] {
            background-color: #ffffff !important;
            border: 1px solid #e2e8f0 !important;
            border-radius: 12px !important;
            box-shadow: 0 8px 30px rgba(15, 23, 42, 0.12) !important;
        }

        /* Dropdown list container */
        ul[data-baseweb="menu"],
        [data-baseweb="menu"] ul {
            background-color: #ffffff !important;
            padding: 4px !important;
        }

        /* Each dropdown option */
        li[role="option"],
        [role="option"] {
            background-color: #ffffff !important;
            color: #1e293b !important;
            border-radius: 8px !important;
        }

        li[role="option"]:hover,
        [role="option"]:hover,
        li[role="option"][aria-selected="true"],
        [role="option"][aria-selected="true"] {
            background-color: #f1f5f9 !important;
            color: #0f172a !important;
        }

        /* Option text */
        li[role="option"] span,
        li[role="option"] div,
        [role="option"] span,
        [role="option"] div {
            color: #1e293b !important;
        }

        /* No-results / empty state text */
        [data-baseweb="menu"] [data-testid="stSelectboxVirtualDropdown"] p,
        [data-baseweb="menu"] p {
            color: #64748b !important;
        }

        /* ─── CHECKBOX & RADIO ────────────────────────────── */
        [data-baseweb="checkbox"] label,
        [data-baseweb="radio"] label {
            color: #1e293b !important;
        }

        /* ─── SLIDER ──────────────────────────────────────── */
        [data-testid="stSlider"] > div > div > div {
            background: #e2e8f0 !important;
        }

        /* ─── TABS ────────────────────────────────────────── */
        [data-baseweb="tab-list"] {
            background: #f1f5f9 !important;
            border-radius: 10px;
        }

        [data-baseweb="tab"] {
            color: #64748b !important;
            background: transparent !important;
        }

        [data-baseweb="tab"][aria-selected="true"] {
            color: #0f172a !important;
            background: #ffffff !important;
            border-radius: 8px;
        }

        [data-baseweb="tab-highlight"] {
            background: #b91c1c !important;
        }

        /* ─── TOOLTIP ─────────────────────────────────────── */
        [data-baseweb="tooltip"],
        [data-baseweb="tooltip"] div {
            background-color: #0f172a !important;
            color: #f1f5f9 !important;
            border-radius: 8px !important;
        }

        /* ─── DATE / TIME PICKER POPUP ────────────────────── */
        [data-baseweb="calendar"],
        [data-baseweb="datepicker"],
        [data-baseweb="time-picker"] {
            background: #ffffff !important;
            border: 1px solid #e2e8f0 !important;
            border-radius: 14px !important;
            box-shadow: 0 8px 30px rgba(15, 23, 42, 0.12) !important;
            color: #1e293b !important;
        }

        [data-baseweb="calendar"] button,
        [data-baseweb="calendar"] div,
        [data-baseweb="calendar"] span {
            color: #1e293b !important;
        }

        /* ─── METRIC CARDS ─────────────────────────────────── */
        div[data-testid="stMetric"] {
            background: #ffffff;
            border: 1px solid #e2e8f0;
            border-radius: 16px;
            padding: 1rem 1.1rem;
            box-shadow: 0 2px 12px rgba(15, 23, 42, 0.06);
        }

        div[data-testid="stMetric"] label {
            color: #64748b !important;
            font-size: 0.78rem !important;
            font-weight: 600 !important;
            text-transform: uppercase;
            letter-spacing: 0.06em;
        }

        div[data-testid="stMetricValue"] > div {
            color: #0f172a !important;
            font-family: 'Space Grotesk', sans-serif !important;
            font-weight: 700 !important;
            font-size: 1.7rem !important;
        }

        /* ─── DATAFRAMES / DATA EDITOR / TABLES ───────────── */
        div[data-testid="stDataFrame"],
        div[data-testid="stDataEditor"] {
            background: #ffffff !important;
            border: 1px solid #e2e8f0 !important;
            border-radius: 14px !important;
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.05) !important;
            overflow: hidden;
            color-scheme: light;
            --background-color: #ffffff;
            --secondary-background-color: #ffffff;
            --text-color: #0f172a;
            --body-text-color: #1e293b;
            --dataframe-header-background-color: #f8fafc;
            --dataframe-border-color: #e2e8f0;
            --dataframe-row-hover-color: #f1f5f9;
        }

        /* Glide grid root used by st.dataframe / st.data_editor */
        div[data-testid="stDataFrame"] [role="grid"],
        div[data-testid="stDataEditor"] [role="grid"],
        div[data-testid="stDataFrame"] .glideDataEditor,
        div[data-testid="stDataEditor"] .glideDataEditor {
            background: #ffffff !important;
            color: #1e293b !important;
        }

        /* Header cells */
        div[data-testid="stDataFrame"] [role="columnheader"],
        div[data-testid="stDataEditor"] [role="columnheader"],
        div[data-testid="stDataFrame"] .gdg-header,
        div[data-testid="stDataEditor"] .gdg-header {
            background: #f8fafc !important;
            color: #334155 !important;
            font-weight: 700 !important;
            border-bottom: 1px solid #e2e8f0 !important;
        }

        /* Body cells */
        div[data-testid="stDataFrame"] [role="gridcell"],
        div[data-testid="stDataEditor"] [role="gridcell"],
        div[data-testid="stDataFrame"] .gdg-cell,
        div[data-testid="stDataEditor"] .gdg-cell {
            background: #ffffff !important;
            color: #1e293b !important;
            border-color: #f1f5f9 !important;
        }

        /* Row hover / selection */
        div[data-testid="stDataFrame"] [role="row"]:hover [role="gridcell"],
        div[data-testid="stDataEditor"] [role="row"]:hover [role="gridcell"],
        div[data-testid="stDataFrame"] [aria-selected="true"],
        div[data-testid="stDataEditor"] [aria-selected="true"] {
            background: #f8fafc !important;
            color: #0f172a !important;
        }

        /* Data editor text inputs inside cells */
        div[data-testid="stDataEditor"] input,
        div[data-testid="stDataEditor"] textarea,
        div[data-testid="stDataEditor"] [contenteditable="true"] {
            background: #ffffff !important;
            color: #0f172a !important;
            caret-color: #b91c1c !important;
        }

        /* Table markdown output (st.table / pandas html) */
        .stTable table,
        .stMarkdown table {
            width: 100%;
            border-collapse: separate;
            border-spacing: 0;
            background: #ffffff;
            color: #1e293b;
            border: 1px solid #e2e8f0;
            border-radius: 12px;
            overflow: hidden;
        }

        .stTable thead th,
        .stMarkdown thead th {
            background: #f8fafc;
            color: #334155;
            font-weight: 700;
            border-bottom: 1px solid #e2e8f0;
        }

        .stTable th,
        .stTable td,
        .stMarkdown th,
        .stMarkdown td {
            padding: 0.55rem 0.75rem;
            border-bottom: 1px solid #f1f5f9;
            color: #1e293b;
        }

        .stTable tbody tr:hover td,
        .stMarkdown tbody tr:hover td {
            background: #f8fafc;
        }

        /* Scrollbar inside table widgets */
        div[data-testid="stDataFrame"] *::-webkit-scrollbar,
        div[data-testid="stDataEditor"] *::-webkit-scrollbar {
            width: 10px;
            height: 10px;
        }

        div[data-testid="stDataFrame"] *::-webkit-scrollbar-track,
        div[data-testid="stDataEditor"] *::-webkit-scrollbar-track {
            background: #f1f5f9;
        }

        div[data-testid="stDataFrame"] *::-webkit-scrollbar-thumb,
        div[data-testid="stDataEditor"] *::-webkit-scrollbar-thumb {
            background: #cbd5e1;
            border-radius: 999px;
            border: 2px solid #f1f5f9;
        }

        /* ─── EXPANDERS ────────────────────────────────────── */
        div[data-testid="stExpander"] {
            background: #ffffff;
            border: 1px solid #e2e8f0 !important;
            border-radius: 14px;
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.05);
            overflow: hidden;
        }

        div[data-testid="stExpander"] summary {
            color: #0f172a !important;
            font-weight: 600;
            background: #ffffff !important;
        }

        div[data-testid="stExpander"] > div > div {
            background: #ffffff !important;
            color: #1e293b !important;
        }

        /* ─── FILE UPLOADER ────────────────────────────────── */
        div[data-testid="stFileUploader"] {
            background: #ffffff !important;
            border-radius: 14px;
        }

        div[data-testid="stFileUploaderDropzone"],
        section[data-testid="stFileUploaderDropzone"] {
            background: #f8fafc !important;
            border: 2px dashed #cbd5e1 !important;
            border-radius: 12px !important;
        }

        div[data-testid="stFileUploaderDropzoneInstructions"],
        div[data-testid="stFileUploaderDropzone"] > div {
            background: transparent !important;
        }

        div[data-testid="stFileUploaderDropzoneInstructions"] span,
        div[data-testid="stFileUploaderDropzoneInstructions"] p,
        div[data-testid="stFileUploaderDropzoneInstructions"] small,
        div[data-testid="stFileUploaderDropzone"] span,
        div[data-testid="stFileUploaderDropzone"] p,
        div[data-testid="stFileUploaderDropzone"] small {
            color: #64748b !important;
        }

        div[data-testid="stFileUploader"] label,
        div[data-testid="stFileUploader"] span,
        div[data-testid="stFileUploader"] p,
        div[data-testid="stFileUploader"] small {
            color: #334155 !important;
        }

        /* Browse files button inside uploader */
        div[data-testid="stFileUploader"] button,
        div[data-testid="stFileUploaderDropzone"] button {
            background: #ffffff !important;
            color: #0f172a !important;
            border: 1.5px solid #e2e8f0 !important;
            border-radius: 8px !important;
        }

        /* ─── ALERTS ───────────────────────────────────────── */
        div[data-testid="stAlert"] {
            border-radius: 12px;
        }

        /* ─── DIVIDER ──────────────────────────────────────── */
        hr {
            border-color: #e2e8f0;
            margin: 1rem 0;
        }

        /* ─── SECTION HEADERS (st.header / st.subheader) ───── */
        .stMarkdown h2, .stMarkdown h3 {
            color: #0f172a !important;
            border-bottom: 2px solid #e2e8f0;
            padding-bottom: 0.35rem;
            margin-top: 1.2rem;
        }

        /* ─── MODAL / DIALOG ──────────────────────────────── */
        [data-testid="stModal"] > div,
        [data-testid="stDialog"] > div {
            background: #ffffff !important;
            color: #0f172a !important;
            border-radius: 18px !important;
        }

        /* ─── PROGRESS BAR ────────────────────────────────── */
        [data-testid="stProgress"] > div > div {
            background: #b91c1c !important;
        }

        [data-testid="stProgress"] > div {
            background: #e2e8f0 !important;
            border-radius: 999px !important;
        }

        /* ─── SIDEBAR EXPANDERS ───────────────────────────── */
        [data-testid="stSidebar"] div[data-testid="stExpander"] {
            background: rgba(255, 255, 255, 0.05) !important;
            border-color: rgba(255, 255, 255, 0.10) !important;
        }

        [data-testid="stSidebar"] div[data-testid="stExpander"] summary {
            color: #cbd5e1 !important;
            background: transparent !important;
        }

        [data-testid="stSidebar"] div[data-testid="stExpander"] > div > div {
            background: transparent !important;
            color: #94a3b8 !important;
        }

        @media (max-width: 900px) {
            .hero-title { font-size: 1.5rem; }
            .block-container { padding-top: 1rem; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

def render_page_header(active_module: str, usuario: str, rol: str):
    meta = MODULE_META.get(active_module, MODULE_META["nuevo"])
    st.markdown(
        f"""
        <div class="hero-card">
            <div class="hero-kicker">Inventarios Rotativos · Grupo Cenoa</div>
            <h1 class="hero-title">{meta['icon']} {meta['title']}</h1>
            <div class="hero-desc">{meta['description']}</div>
            <div class="hero-meta">
                <span class="hero-pill">Usuario: {usuario}</span>
                <span class="hero-pill">Rol: {rol}</span>
                <span class="hero-pill">Backend: {DB_BACKEND}</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

# ----------------------------
# DATABASE FUNCTIONS
# ----------------------------
@st.cache_resource
def get_db_engine():
    connect_args = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
    return create_engine(DATABASE_URL, future=True, connect_args=connect_args)

def init_database():
    try:
        engine = get_db_engine()
        with engine.begin() as conn:
            if DATABASE_URL.startswith("sqlite"):
                conn.execute(text("PRAGMA journal_mode=WAL"))
                conn.execute(text("PRAGMA synchronous=NORMAL"))
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS worksheet_store (
                    name TEXT PRIMARY KEY,
                    data_json TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
            """))
    except Exception as e:
        st.error(f"Error inicializando base de datos: {e}")
        st.stop()

init_database()

@st.cache_data(ttl=5)
def read_gspread_worksheet(ws_name: str) -> pd.DataFrame:
    """Read logical worksheet from configured database."""
    try:
        engine = get_db_engine()
        with engine.begin() as conn:
            row = conn.execute(
                text("SELECT data_json FROM worksheet_store WHERE name = :name"),
                {"name": ws_name}
            ).fetchone()
        if not row:
            return pd.DataFrame()
        data = json.loads(row[0]) if row[0] else []
        return pd.DataFrame(data) if data else pd.DataFrame()
    except Exception as e:
        st.error(f"Error reading {ws_name}: {e}")
        return pd.DataFrame()

def write_gspread_worksheet(ws_name: str, df: pd.DataFrame):
    """Write logical worksheet to configured database. Returns (ok: bool, message: str)."""
    try:
        df = df.copy()
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].astype(str)

        df = df.where(pd.notnull(df), "")
        df = df.replace([np.inf, -np.inf], "")

        payload = json.dumps(df.to_dict(orient="records"), ensure_ascii=False, default=str)
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        engine = get_db_engine()
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM worksheet_store WHERE name = :name"), {"name": ws_name})
            conn.execute(
                text("INSERT INTO worksheet_store(name, data_json, updated_at) VALUES (:name, :data_json, :updated_at)"),
                {"name": ws_name, "data_json": payload, "updated_at": now}
            )

        try:
            st.cache_data.clear()
        except Exception:
            pass
        return True, ""
    except Exception as e:
        user_msg = f"Error writing {ws_name}: {e}"
        try:
            st.error(user_msg)
        except Exception:
            pass
        return False, user_msg

def append_gspread_worksheet(ws_name: str, df_new: pd.DataFrame):
    """Append rows to logical worksheet in configured database."""
    try:
        df_new = df_new.copy()
        for col in df_new.columns:
            if pd.api.types.is_datetime64_any_dtype(df_new[col]):
                df_new[col] = df_new[col].astype(str)

        df_exist = read_gspread_worksheet(ws_name)
        if df_exist.empty:
            ok, msg = write_gspread_worksheet(ws_name, df_new)
            if not ok:
                st.error(f"Append failed writing new sheet {ws_name}: {msg}")
            return bool(ok)

        for col in df_exist.columns:
            if col not in df_new.columns:
                df_new[col] = ""
        for col in df_new.columns:
            if col not in df_exist.columns:
                df_exist[col] = ""

        df_final = pd.concat([df_exist, df_new[df_exist.columns]], ignore_index=True)
        ok, msg = write_gspread_worksheet(ws_name, df_final)
        if not ok:
            st.error(f"Append failed updating {ws_name}: {msg}")
        return bool(ok)
    except Exception as e:
        st.error(f"Error appending to {ws_name}: {str(e)}")
        return False


def log_audit(action: str, id_inv: str, filas: int, status: str, mensaje: str = ""):
    """Append an audit row to the Audit_Log sheet. Non-blocking: failures are logged to UI but do not raise."""
    try:
        row = pd.DataFrame([{
            "Timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Usuario": st.session_state.get("usuario", ""),
            "Rol": st.session_state.get("rol", ""),
            "Accion": action,
            "ID_Inventario": id_inv,
            "Filas": int(filas) if filas is not None else 0,
            "Status": status,
            "Mensaje": mensaje
        }])
        append_gspread_worksheet(SHEET_AUDIT, row)
    except Exception as e:
        # Non-fatal: show a warning in the UI for admin visibility
        try:
            st.warning(f"No se pudo escribir Audit_Log: {e}")
        except Exception:
            pass

# ----------------------------
# EXPORT FUNCTIONS
# ----------------------------
def export_dataframe_to_excel(df: pd.DataFrame, sheet_name: str = "Datos", title: str = None) -> io.BytesIO:
    """Export DataFrame to Excel with Argentine number formatting (. for thousands, , for decimals)"""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name limit is 31 chars
    
    # Add title if provided
    if title:
        ws.merge_cells("A1:Z1")
        ws["A1"] = title
        ws["A1"].font = Font(size=12, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        start_row = 3
    else:
        start_row = 1
    
    # Track numeric columns for formatting
    numeric_cols = {}
    for col_idx, col_name in enumerate(df.columns, start=1):
        # Try to detect numeric columns
        try:
            numeric_test = pd.to_numeric(df[col_name], errors="coerce")
            if numeric_test.notna().sum() > 0:  # At least one numeric value
                numeric_cols[col_idx] = col_name
        except:
            pass
    
    # Write DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = Font(bold=True)
            
            # Apply numeric formatting if this column is numeric
            if c_idx in numeric_cols:
                try:
                    num_val = float(value) if value not in (None, "", "NaN", "nan") else None
                    if num_val is not None and str(num_val) not in ("nan", "inf", "-inf"):
                        cell.value = num_val
                        # Argentine format: #,##0.00 with . as thousands separator and , as decimal
                        cell.number_format = '#,##0.00'
                except:
                    pass
            
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def parse_ar_number(series: pd.Series) -> pd.Series:
    """Parse numbers that may use Argentine formatting (1.234,56)."""
    s = series.astype(str).str.strip()
    has_comma = s.str.contains(",", regex=False)
    s = s.where(~has_comma, s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    return pd.to_numeric(s, errors="coerce")

def normalize_article_code(value) -> str:
    if pd.isna(value):
        return ""
    code = str(value).strip()
    if not code:
        return ""
    if code.endswith(".0"):
        base = code[:-2]
        if base.isdigit():
            return base
    return code

def format_number_ar(value, decimals: int = 2) -> str:
    number = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(number):
        return ""
    formatted = f"{float(number):,.{decimals}f}"
    return formatted.replace(",", "_").replace(".", ",").replace("_", ".")

def format_currency_ar(value) -> str:
    formatted = format_number_ar(value, decimals=2)
    return f"$ {formatted}" if formatted else ""

def buscar_articulo_en_base(id_inv: str, codigo_articulo: str) -> dict | None:
    codigo = normalize_article_code(codigo_articulo)
    if not codigo:
        return None

    df_base = read_gspread_worksheet(SHEET_BASE)
    if df_base.empty or "ID_Inventario" not in df_base.columns or C_ART not in df_base.columns:
        return None

    df_inv = df_base[df_base["ID_Inventario"].astype(str) == str(id_inv)].copy()
    if df_inv.empty:
        return None

    codigos = df_inv[C_ART].apply(normalize_article_code)
    matches = df_inv[codigos == codigo].copy()
    if matches.empty:
        return None

    descripcion = matches[C_DESC].dropna().astype(str).iloc[0] if C_DESC in matches.columns and matches[C_DESC].notna().any() else ""
    costo = parse_ar_number(matches[C_COSTO]).dropna().iloc[0] if C_COSTO in matches.columns and parse_ar_number(matches[C_COSTO]).notna().any() else 0.0
    stock_total = parse_ar_number(matches[C_STOCK]).fillna(0).sum() if C_STOCK in matches.columns else 0.0

    return {
        "codigo": codigo,
        "descripcion": descripcion,
        "costo": float(costo) if pd.notna(costo) else 0.0,
        "stock": float(stock_total),
    }

def is_currency_column(col_name: str) -> bool:
    name = str(col_name).strip().lower()
    currency_hints = (
        "costo",
        "precio",
        "cto.rep",
        "cto rep",
        "valor",
        "valuación",
        "valuacion",
        "$",
    )
    return any(hint in name for hint in currency_hints)

def prepare_currency_display(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    if df is None or df.empty:
        return df, {}

    df_view = df.copy()

    for col in df_view.columns:
        if is_currency_column(col):
            df_view[col] = parse_ar_number(df_view[col]).apply(format_currency_ar)

    return df_view, {}

def render_dataframe(df: pd.DataFrame, column_config: dict | None = None, **kwargs):
    df_view, auto_currency_config = prepare_currency_display(df)
    merged_config = {**auto_currency_config, **(column_config or {})}
    if merged_config:
        return st.dataframe(df_view, column_config=merged_config, **kwargs)
    return st.dataframe(df_view, **kwargs)

def render_data_editor(df: pd.DataFrame, column_config: dict | None = None, **kwargs):
    df_view, auto_currency_config = prepare_currency_display(df)
    merged_config = {**auto_currency_config, **(column_config or {})}
    if merged_config:
        return st.data_editor(df_view, column_config=merged_config, **kwargs)
    return st.data_editor(df_view, **kwargs)

# ----------------------------
def verify_password(password: str, password_hash: str) -> bool:
    """Verify password against bcrypt hash"""
    return bcrypt.checkpw(password.encode(), password_hash.encode())

def inject_login_theme():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;600;700&family=Inter:wght@400;500;600&display=swap');

        html, body, [class*="css"] {
            font-family: 'Inter', sans-serif;
        }

        /* Full-screen dark gradient background */
        .stApp {
            background: linear-gradient(135deg, #0c1220 0%, #111827 55%, #0f1f38 100%);
        }

        [data-testid="stAppViewContainer"] > .main {
            background: transparent;
        }

        .block-container {
            padding-top: 2.5rem;
            padding-bottom: 2rem;
            max-width: 1200px;
        }

        /* Hide default header bar & footer */
        #MainMenu { visibility: hidden; }
        footer { visibility: hidden; }
        header { visibility: hidden; }

        /* All text light on dark background */
        p, span, div, label, small {
            color: #cbd5e1;
        }

        h1, h2, h3 {
            font-family: 'Space Grotesk', sans-serif !important;
            color: #f1f5f9 !important;
        }

        /* Login card (the form container) */
        div[data-testid="stForm"] {
            background: rgba(255, 255, 255, 0.04);
            border: 1px solid rgba(255, 255, 255, 0.10);
            border-radius: 22px;
            padding: 0.5rem 0.5rem 0.5rem 0.5rem;
        }

        /* Input fields */
        div[data-testid="stTextInputRootElement"] > div {
            background: #ffffff !important;
            border: 1px solid rgba(148, 163, 184, 0.45) !important;
            border-radius: 10px !important;
        }

        div[data-testid="stTextInputRootElement"] input {
            color: #0f172a !important;
            -webkit-text-fill-color: #0f172a !important;
            caret-color: #b91c1c !important;
            background: #ffffff !important;
        }

        div[data-testid="stTextInputRootElement"] input::placeholder {
            color: #64748b !important;
            -webkit-text-fill-color: #64748b !important;
        }

        /* Input labels */
        div[data-testid="stTextInputRootElement"] label,
        .stTextInput label {
            color: #94a3b8 !important;
            font-size: 0.8rem !important;
            font-weight: 600 !important;
            text-transform: uppercase;
            letter-spacing: 0.08em;
        }

        /* Submit button */
        div[data-testid="stFormSubmitButton"] > button {
            background: linear-gradient(135deg, #b91c1c 0%, #dc2626 100%) !important;
            color: #ffffff !important;
            border: none !important;
            border-radius: 12px !important;
            font-family: 'Space Grotesk', sans-serif !important;
            font-weight: 700 !important;
            font-size: 1rem !important;
            letter-spacing: 0.02em;
            padding: 0.75rem 0 !important;
            box-shadow: 0 6px 28px rgba(185, 28, 28, 0.45) !important;
            transition: all 0.2s ease !important;
        }

        div[data-testid="stFormSubmitButton"] > button:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 10px 36px rgba(185, 28, 28, 0.55) !important;
        }

        /* Credentials expander */
        div[data-testid="stExpander"] {
            background: rgba(255, 255, 255, 0.03);
            border: 1px solid rgba(255, 255, 255, 0.08);
            border-radius: 14px;
        }

        div[data-testid="stExpander"] summary {
            color: #94a3b8 !important;
            font-size: 0.85rem;
        }

        div[data-testid="stDataFrame"] {
            background: rgba(255, 255, 255, 0.04);
            border: 1px solid rgba(255, 255, 255, 0.08);
            border-radius: 10px;
        }

        /* Alerts */
        div[data-testid="stAlert"] {
            border-radius: 12px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

def login():
    inject_login_theme()

    _, col_center, _ = st.columns([1, 2, 1])

    with col_center:
        # Brand header
        st.markdown(
            """
            <div style="text-align:center; margin-bottom: 2rem; padding-top: 0.5rem;">
                <div style="font-size: 2.8rem; margin-bottom: 0.6rem;">📦</div>
                <div style="
                    font-size: 0.7rem;
                    font-weight: 700;
                    text-transform: uppercase;
                    letter-spacing: 0.18em;
                    color: #64748b;
                    margin-bottom: 0.5rem;
                ">AUDITORÍA INTERNA · GRUPO CENOA</div>
                <div style="
                    font-family: 'Space Grotesk', sans-serif;
                    font-size: 1.75rem;
                    font-weight: 700;
                    color: #f1f5f9;
                    line-height: 1.2;
                ">Inventarios Rotativos</div>
                <div style="
                    font-size: 0.9rem;
                    color: #475569;
                    margin-top: 0.4rem;
                ">Acceso restringido al personal autorizado</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        with st.form("login_form"):
            usuario = st.text_input("Usuario (ID)", placeholder="Ej: diego_guantay")
            contrasena = st.text_input("Contraseña", type="password", placeholder="••••••••")
            st.write("")
            submit = st.form_submit_button("Ingresar →", use_container_width=True)

            if submit:
                if usuario in USUARIOS_CREDENCIALES:
                    creds = USUARIOS_CREDENCIALES[usuario]
                    if verify_password(contrasena, creds["password_hash"]):
                        st.session_state["logged_in"] = True
                        st.session_state["usuario"] = usuario
                        st.session_state["nombre_usuario"] = creds["nombre"]
                        st.session_state["rol"] = creds["rol"]
                        st.success(f"✅ Bienvenido, {creds['nombre']}!")
                        st.rerun()
                    else:
                        st.error("Contraseña incorrecta")
                else:
                    st.error("Usuario no encontrado")

        st.write("")
        with st.expander("📋 Credenciales de prueba (eliminar en producción)"):
            creds_data = []
            for user_id, password in CREDENCIALES_INICIALES.items():
                creds_data.append({
                    "Usuario (ID)": user_id,
                    "Contraseña": password,
                    "Rol": USUARIOS_CREDENCIALES[user_id]["rol"],
                    "Nombre": USUARIOS_CREDENCIALES[user_id]["nombre"],
                })
            render_dataframe(pd.DataFrame(creds_data), use_container_width=True, hide_index=True)
            st.caption("⚠️ Eliminar antes de pasar a producción.")

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    login()
    st.stop()

inject_modern_theme()

usuario_actual = st.session_state.get("usuario")
nombre_actual = st.session_state.get("nombre_usuario")
rol_actual = st.session_state.get("rol")

# --- Admin debug: mostrar estado de datos (solo para admin)
def _admin_debug_show():
    try:
        dfh = read_gspread_worksheet(SHEET_HIST)
        dfd = read_gspread_worksheet(SHEET_DET)
    except Exception as e:
        st.sidebar.error(f"Debug read error: {e}")
        return

    with st.sidebar.expander("Base de datos (admin)", expanded=False):
        st.write("**Historial_Inventarios**")
        st.write("Rows:", 0 if dfh is None else len(dfh))
        st.write("Columns:", list(dfh.columns) if not (dfh is None or dfh.empty) else [])
        st.write("---")
        st.write("**Detalle_Articulos**")
        st.write("Rows:", 0 if dfd is None else len(dfd))
        st.write("Columns:", list(dfd.columns) if not (dfd is None or dfd.empty) else [])
        st.write("---")
        st.info("Este panel muestra solo conteos y nombres de columnas para depuración.")

        # Mostrar últimos registros de Audit_Log si existe
        try:
            dfa = read_gspread_worksheet(SHEET_AUDIT)
            if not dfa.empty:
                st.write("**Audit_Log (últimas 10 filas)**")
                render_dataframe(dfa.tail(10).sort_values("Timestamp", ascending=False), use_container_width=True)
            else:
                st.write("**Audit_Log**: (vacío)")
        except Exception as e:
            st.write(f"Audit_Log: error al leer: {e}")

if usuario_actual == "admin":
    _admin_debug_show()
    with st.sidebar.expander("Configuración BD", expanded=False):
        st.write("Backend activo:")
        st.write(DB_BACKEND)
        st.write("Destino:")
        st.write(str(DB_PATH) if DB_BACKEND == "SQLite" else "DATABASE_URL configurada")

# ----------------------------
# DATA FUNCTIONS
# ----------------------------
def listar_inventarios_abiertos():
    df_hist = read_gspread_worksheet(SHEET_HIST)
    if df_hist.empty or "Estado" not in df_hist.columns:
        return pd.DataFrame()
    return df_hist[df_hist["Estado"].astype(str).str.lower() == "abierto"].copy()

def cargar_detalle(id_inv: str) -> pd.DataFrame:
    df = read_gspread_worksheet(SHEET_DET)
    if df.empty or "ID_Inventario" not in df.columns:
        return pd.DataFrame()
    return df[df["ID_Inventario"].astype(str) == str(id_inv)].copy()

def calcular_resultados_inventario(df_det: pd.DataFrame) -> dict:
    """Calculate inventory results based on Auditor adjustments.
    - Muestra Q & $ Ajuste: sum of ALL articles in sample (no filter)
    - Faltantes/Sobrantes/etc: only rows where Tipo_Ajuste = "Ajuste"
    - %: always over total muestra valuation
    """
    if df_det.empty:
        return {"canjes": []}
    
    df_all = df_det.copy()
    stock_col = C_STOCK if C_STOCK in df_all.columns else None
    costo_col = C_COSTO if C_COSTO in df_all.columns else None
    
    if not stock_col or not costo_col:
        return {"canjes": []}
    
    df_all["_stock"] = parse_ar_number(df_all[stock_col]).fillna(0)
    df_all["_costo"] = parse_ar_number(df_all[costo_col]).fillna(0)
    
    # MUESTRA: sum of ALL articles (no filter)
    cant_muestra = int(df_all["_stock"].sum())
    valor_muestra = (df_all["_stock"] * df_all["_costo"]).sum()
    pct_muestra = 100.0
    
    # Now filter to ONLY "Ajuste" rows for difference calculations
    df_r = df_all.copy()
    if "Tipo_Ajuste" in df_r.columns:
        mask_ajuste = df_r["Tipo_Ajuste"].astype(str) == "Ajuste"
        df_r = df_r[mask_ajuste].copy()
    
    # If no adjustments, return results with 0 differences
    if df_r.empty:
        return {
            "cant_muestra": cant_muestra,
            "valor_muestra": valor_muestra,
            "pct_muestra": pct_muestra,
            "cant_faltantes": 0,
            "valor_faltantes": 0,
            "pct_faltantes": 0,
            "cant_sobrantes": 0,
            "valor_sobrantes": 0,
            "pct_sobrantes": 0,
            "cant_dif_neta": 0,
            "valor_dif_neta": 0,
            "pct_dif_neta": 0,
            "cant_dif_absoluta": 0,
            "valor_dif_absoluta": 0,
            "pct_dif_absoluta": 0,
            "pct_absoluto": 0,
            "grado": 100,
            "escala": [(0.00, 100), (0.10, 94), (0.80, 82), (1.60, 65), (2.40, 35), (3.30, 0)],
            "canjes": []
        }
    
    ajuste_source = df_r["Ajuste_Cantidad"] if "Ajuste_Cantidad" in df_r.columns else pd.Series(0, index=df_r.index)
    df_r["_ajuste"] = pd.to_numeric(ajuste_source, errors="coerce").fillna(0)
    
    # Faltantes (negative adjustments)
    mask_falt = df_r["_ajuste"] < 0
    cant_faltantes = int((df_r.loc[mask_falt, "_ajuste"].abs()).sum())
    valor_faltantes = (df_r.loc[mask_falt, "_ajuste"].abs() * df_r.loc[mask_falt, "_costo"]).sum()
    pct_faltantes = (valor_faltantes / valor_muestra * 100) if valor_muestra > 0 else 0
    
    # Sobrantes (positive adjustments)
    mask_sobr = df_r["_ajuste"] > 0
    cant_sobrantes = int(df_r.loc[mask_sobr, "_ajuste"].sum())
    valor_sobrantes = (df_r.loc[mask_sobr, "_ajuste"] * df_r.loc[mask_sobr, "_costo"]).sum()
    pct_sobrantes = (valor_sobrantes / valor_muestra * 100) if valor_muestra > 0 else 0
    
    # Diferencia neta y absoluta
    cant_dif_neta = int(df_r["_ajuste"].sum())
    valor_dif_neta = (df_r["_ajuste"] * df_r["_costo"]).sum()
    pct_dif_neta = (valor_dif_neta / valor_muestra * 100) if valor_muestra > 0 else 0
    
    cant_dif_absoluta = int(df_r["_ajuste"].abs().sum())
    valor_dif_absoluta = (df_r["_ajuste"].abs() * df_r["_costo"]).sum()
    pct_dif_absoluta = (valor_dif_absoluta / valor_muestra * 100) if valor_muestra > 0 else 0
    
    # Escala de grado basada en % absoluto
    pct_absoluto = pct_dif_absoluta
    escala = [(0.00, 100), (0.10, 94), (0.80, 82), (1.60, 65), (2.40, 35), (3.30, 0)]
    escala_sorted = sorted(escala, key=lambda x: x[0])
    grado = 0
    for th, g in escala_sorted:
        if pct_absoluto >= th:
            grado = g
    
    # Collect canjes (separate from adjustments)
    canjes_list = []
    if "Tipo_Ajuste" in df_det.columns:
        df_canjes = df_det[df_det["Tipo_Ajuste"].astype(str) == "Canje"].copy()
        if not df_canjes.empty:
            for idx, row in df_canjes.iterrows():
                art = row.get("Canje_Articulo", "") or row.get(C_ART, "")
                descripcion = row.get("Canje_Descripcion", "")
                loc = row.get(C_LOC, "")
                costo = pd.to_numeric(row.get("Canje_Costo_Rep", row.get(C_COSTO, 0)), errors="coerce")
                stock_base = pd.to_numeric(row.get("Canje_Stock_Base", 0), errors="coerce")
                ajuste_cant = pd.to_numeric(row.get("Ajuste_Cantidad", 0), errors="coerce")
                canjes_list.append({
                    "Artículo": art,
                    "Descripción": descripcion,
                    "Locación": loc,
                    "Stock Base": stock_base,
                    "Cantidad": ajuste_cant,
                    "Costo Unitario": costo,
                    "Valor Total": ajuste_cant * costo
                })
    
    return {
        "cant_muestra": cant_muestra,
        "valor_muestra": valor_muestra,
        "pct_muestra": pct_muestra,
        "cant_faltantes": cant_faltantes,
        "valor_faltantes": valor_faltantes,
        "pct_faltantes": pct_faltantes,
        "cant_sobrantes": cant_sobrantes,
        "valor_sobrantes": valor_sobrantes,
        "pct_sobrantes": pct_sobrantes,
        "cant_dif_neta": cant_dif_neta,
        "valor_dif_neta": valor_dif_neta,
        "pct_dif_neta": pct_dif_neta,
        "cant_dif_absoluta": cant_dif_absoluta,
        "valor_dif_absoluta": valor_dif_absoluta,
        "pct_dif_absoluta": pct_dif_absoluta,
        "pct_absoluto": pct_absoluto,
        "grado": grado,
        "escala": escala_sorted,
        "canjes": canjes_list
    }

def guardar_detalle_modificado(id_inv: str, df_mod: pd.DataFrame):
    """Update inventory details"""
    try:
        df_all = read_gspread_worksheet(SHEET_DET)
        if df_all.empty:
            ok, msg = write_gspread_worksheet(SHEET_DET, df_mod)
            log_audit("guardar_detalle", id_inv, len(df_mod), "OK" if ok else "ERROR", msg if msg else "Creó hoja o sobreescribió")
            return bool(ok)

        df_all = df_all.copy()
        mask = df_all["ID_Inventario"].astype(str) == str(id_inv)
        df_rest = df_all.loc[~mask].copy()
        df_final = pd.concat([df_rest, df_mod], ignore_index=True)
        ok, msg = write_gspread_worksheet(SHEET_DET, df_final)
        log_audit("guardar_detalle", id_inv, len(df_mod), "OK" if ok else "ERROR", msg if msg else "Actualizó detalle")
        return bool(ok)
    except Exception as e:
        log_audit("guardar_detalle", id_inv, 0, "ERROR", str(e))
        return False

def cerrar_inventario(id_inv: str, usuario: str):
    """Close inventory"""
    df_hist = read_gspread_worksheet(SHEET_HIST)
    if df_hist.empty or "ID_Inventario" not in df_hist.columns:
        return
    df_hist = df_hist.copy()
    mask = df_hist["ID_Inventario"].astype(str) == str(id_inv)
    if mask.sum() == 0:
        return
    df_hist.loc[mask, "Estado"] = "Cerrado"
    df_hist.loc[mask, "Cierre_Fecha"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    df_hist.loc[mask, "Cierre_Usuario"] = usuario
    ok, msg = write_gspread_worksheet(SHEET_HIST, df_hist)
    log_audit("cerrar_inventario", id_inv, 0, "OK" if ok else "ERROR", msg if msg else "Cerró inventario")

def calcular_dashboard_kpis() -> dict:
    df_hist = read_gspread_worksheet(SHEET_HIST)
    df_det = read_gspread_worksheet(SHEET_DET)

    if df_hist.empty:
        return {
            "inventarios_totales": 0,
            "inventarios_abiertos": 0,
            "inventarios_cerrados": 0,
            "tasa_cierre": 0.0,
            "lineas_muestreadas": 0,
            "valuacion_muestra": 0.0,
            "exactitud_promedio": 0.0,
            "detalle_resumen": pd.DataFrame(),
            "ranking_sucursales": pd.DataFrame(),
        }

    df_hist = df_hist.copy()
    estados = df_hist.get("Estado", pd.Series(dtype=str)).astype(str).str.strip().str.lower()
    inventarios_totales = len(df_hist)
    inventarios_abiertos = int((estados == "abierto").sum())
    inventarios_cerrados = int((estados == "cerrado").sum())
    tasa_cierre = (inventarios_cerrados / inventarios_totales * 100) if inventarios_totales else 0.0

    lineas_muestreadas = len(df_det) if not df_det.empty else 0
    valuacion_muestra = 0.0
    exactitudes = []
    resumen_inventarios = []

    if not df_det.empty and "ID_Inventario" in df_det.columns:
        df_det = df_det.copy()
        df_det["_stock"] = parse_ar_number(df_det.get(C_STOCK, pd.Series(dtype=object))).fillna(0)
        df_det["_costo"] = parse_ar_number(df_det.get(C_COSTO, pd.Series(dtype=object))).fillna(0)
        df_det["_diferencia"] = parse_ar_number(df_det.get("Diferencia", pd.Series(dtype=object))).fillna(0)
        df_det["_valor_linea"] = df_det["_stock"] * df_det["_costo"]
        valuacion_muestra = float(df_det["_valor_linea"].sum())

        for id_inv, grupo in df_det.groupby("ID_Inventario"):
            resultados = calcular_resultados_inventario(grupo.copy())
            exactitudes.append(float(resultados.get("grado", 0)))
            resumen_inventarios.append({
                "ID_Inventario": id_inv,
                "Sucursal": grupo.get("Sucursal", pd.Series([""])).iloc[0] if not grupo.empty else "",
                "Líneas": len(grupo),
                "Valuación": float((grupo["_stock"] * grupo["_costo"]).sum()),
                "Líneas con diferencia": int((grupo["_diferencia"] != 0).sum()),
                "Exactitud": float(resultados.get("grado", 0)),
            })

        detalle_resumen = pd.DataFrame(resumen_inventarios).sort_values(["Exactitud", "Valuación"], ascending=[True, False])

        ranking_sucursales = (
            pd.DataFrame(resumen_inventarios)
            .groupby("Sucursal", dropna=False)
            .agg(
                Inventarios=("ID_Inventario", "count"),
                Valuación=("Valuación", "sum"),
                Exactitud_Promedio=("Exactitud", "mean"),
            )
            .reset_index()
            .sort_values(["Exactitud_Promedio", "Valuación"], ascending=[True, False])
        ) if resumen_inventarios else pd.DataFrame()
    else:
        detalle_resumen = pd.DataFrame()
        ranking_sucursales = pd.DataFrame()

    exactitud_promedio = float(np.mean(exactitudes)) if exactitudes else 0.0

    return {
        "inventarios_totales": inventarios_totales,
        "inventarios_abiertos": inventarios_abiertos,
        "inventarios_cerrados": inventarios_cerrados,
        "tasa_cierre": tasa_cierre,
        "lineas_muestreadas": lineas_muestreadas,
        "valuacion_muestra": valuacion_muestra,
        "exactitud_promedio": exactitud_promedio,
        "detalle_resumen": detalle_resumen,
        "ranking_sucursales": ranking_sucursales,
    }

# ----------------------------
# UI
# ----------------------------
with st.sidebar:
    st.markdown(
        """
        <div class="brand-shell">
            <div class="brand-kicker">Auditoría Interna</div>
            <div class="brand-title">Inventarios Rotativos<br/>Grupo Cenoa</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.write("### 📌 Módulos")

    if "modulo_activo" not in st.session_state:
        st.session_state["modulo_activo"] = "nuevo"

    if st.button(f"{MODULE_META['nuevo']['icon']}  {MODULE_META['nuevo']['label']}", use_container_width=True):
        st.session_state["modulo_activo"] = "nuevo"
    if st.button(f"{MODULE_META['conteo']['icon']}  {MODULE_META['conteo']['label']}", use_container_width=True):
        st.session_state["modulo_activo"] = "conteo"
    if st.button(f"{MODULE_META['justificaciones']['icon']}  {MODULE_META['justificaciones']['label']}", use_container_width=True):
        st.session_state["modulo_activo"] = "justificaciones"
    if st.button(f"{MODULE_META['cierre']['icon']}  {MODULE_META['cierre']['label']}", use_container_width=True):
        st.session_state["modulo_activo"] = "cierre"
    if st.button(f"{MODULE_META['dashboards']['icon']}  {MODULE_META['dashboards']['label']}", use_container_width=True):
        st.session_state["modulo_activo"] = "dashboards"

    st.write("---")
    st.write(f"**👤 Logueado como:** {nombre_actual}")
    st.write(f"**🎯 Rol:** {rol_actual}")
    
    if st.button("🚪 Cerrar sesión", use_container_width=True):
        st.session_state["logged_in"] = False
        st.session_state.clear()
        st.rerun()

modulo_activo = st.session_state.get("modulo_activo", "nuevo")
render_page_header(modulo_activo, nombre_actual, rol_actual)

# ----------------------------
# MÓDULO 1
# ----------------------------
if modulo_activo == "nuevo":
    st.subheader("Panel de control del Auditor")

    c1, c2 = st.columns(2)
    with c1:
        concesionaria = st.selectbox("Concesionaria", list(CONCESIONARIAS.keys()))
    with c2:
        sucursal = st.selectbox("Sucursal", CONCESIONARIAS[concesionaria])

    st.divider()

    if rol_actual not in ("Auditor", "admin"):
        st.info("Solo Auditores pueden generar inventarios.")
    else:
        st.subheader("Importar Excel → ABC → Muestra 80/15/5")

        archivo = st.file_uploader("Subir reporte de stock (.xlsx)", type=["xlsx"])

        if archivo:
            df_base = pd.read_excel(archivo)
            st.write("Vista previa:")
            render_dataframe(df_base.head(15), use_container_width=True)

            if st.button("✅ Generar y guardar inventario"):
                falt = [c for c in [C_ART, C_LOC, C_DESC, C_STOCK, C_COSTO] if c not in df_base.columns]
                if falt:
                    st.error(f"Faltan columnas: {', '.join(falt)}")
                    st.stop()

                df = df_base.copy()
                df[C_STOCK] = pd.to_numeric(df[C_STOCK], errors="coerce").fillna(0)
                df[C_COSTO] = pd.to_numeric(df[C_COSTO], errors="coerce").fillna(0)

                df["Valor_T"] = df[C_STOCK] * df[C_COSTO]
                total = df["Valor_T"].sum()
                if total <= 0:
                    st.error("No se puede calcular ABC")
                    st.stop()

                df = df.sort_values("Valor_T", ascending=False)
                df["Acc"] = df["Valor_T"].cumsum() / total
                df["Cat"] = df["Acc"].apply(lambda x: "A" if x <= 0.8 else ("B" if x <= 0.95 else "C"))

                df_a = df[df["Cat"] == "A"]
                df_b = df[df["Cat"] == "B"]
                df_c = df[df["Cat"] == "C"]

                m_a = df_a.sample(n=min(80, len(df_a))) if len(df_a) else df_a
                m_b = df_b.sample(n=min(15, len(df_b))) if len(df_b) else df_b
                m_c = df_c.sample(n=min(5, len(df_c))) if len(df_c) else df_c

                muestra = pd.concat([m_a, m_b, m_c], ignore_index=True)

                muestra["Concesionaria"] = concesionaria
                muestra["Sucursal"] = sucursal
                muestra["Conteo_Fisico"] = ""
                muestra["Diferencia"] = ""
                muestra["Justificacion"] = ""
                muestra["Justif_Validada"] = ""
                muestra["Validador"] = ""
                muestra["Fecha_Validacion"] = ""

                id_inv = datetime.datetime.now().strftime("INV-%Y%m%d-%H%M")

                df_hist = read_gspread_worksheet(SHEET_HIST)
                if not df_hist.empty and "ID_Inventario" in df_hist.columns:
                    if (df_hist["ID_Inventario"].astype(str) == id_inv).any():
                        st.warning("ID ya existe")
                        st.stop()

                nueva_fila = pd.DataFrame([{
                    "ID_Inventario": id_inv,
                    "Fecha": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "Concesionaria": concesionaria,
                    "Sucursal": sucursal,
                    "Auditor": usuario_actual,
                    "Estado": "Abierto",
                    "Cierre_Fecha": "",
                    "Cierre_Usuario": ""
                }])
                ok_hist = append_gspread_worksheet(SHEET_HIST, nueva_fila)

                df_base_store = df_base.copy()
                if C_ART in df_base_store.columns:
                    df_base_store[C_ART] = df_base_store[C_ART].apply(normalize_article_code)
                df_base_store["ID_Inventario"] = id_inv
                df_base_store["Concesionaria"] = concesionaria
                df_base_store["Sucursal"] = sucursal
                ok_base = append_gspread_worksheet(SHEET_BASE, df_base_store)

                muestra["ID_Inventario"] = id_inv
                ok_det = append_gspread_worksheet(SHEET_DET, muestra)

                # Log actions
                log_audit("generar_inventario", id_inv, len(muestra), "OK" if (ok_hist and ok_det and ok_base) else "ERROR", f"hist_ok={ok_hist}, det_ok={ok_det}, base_ok={ok_base}")

                if ok_hist and ok_det and ok_base:
                    st.success(f"✅ Inventario {id_inv} creado y detalle guardado ({len(muestra)} filas).")
                elif ok_hist and ok_det and not ok_base:
                    st.warning(f"Inventario {id_inv} creado, pero no se pudo guardar la base completa del Excel. La búsqueda para canjes puede fallar.")
                elif ok_hist and not ok_det:
                    st.warning(f"Inventario {id_inv} creado en historial, pero no se pudo guardar el detalle.")
                else:
                    st.error("No se pudo crear el inventario. Revisá los mensajes de error.")

                # Mostrar confirmación / chequeo rápido del detalle (solo conteos y columnas)
                try:
                    df_det_check = read_gspread_worksheet(SHEET_DET)
                    if not df_det_check.empty and "ID_Inventario" in df_det_check.columns:
                        cnt = int((df_det_check["ID_Inventario"].astype(str) == str(id_inv)).sum())
                        st.info(f"Detalle guardado: {cnt} filas para {id_inv} (total hoja: {len(df_det_check)}).")
                    else:
                        st.info("Detalle no encontrado o estructura no contiene 'ID_Inventario'.")
                except Exception as e:
                    st.info(f"Chequeo detalle: error al leer hoja: {e}")

                # Opción de descargar muestra generada
                st.divider()
                st.write("### 📥 Descargar muestra:")
                cols_export = [C_ART, C_LOC, C_DESC, C_STOCK, C_COSTO, "Cat", "Concesionaria", "Sucursal"]
                cols_export = [c for c in cols_export if c in muestra.columns]
                df_export = muestra[cols_export].copy()
                xlsx_data = export_dataframe_to_excel(df_export, sheet_name="Muestra", title=f"Muestra Inventario {id_inv}")
                st.download_button(
                    "⬇️ Descargar Muestra Excel",
                    data=xlsx_data,
                    file_name=f"Muestra_{id_inv}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                # Refrescar vista (el usuario puede volver a abrir la pestaña o recargar)
                st.rerun()

# ----------------------------
# MÓDULO 2
# ----------------------------
elif modulo_activo == "conteo":
    st.subheader("Carga de conteo físico")

    if rol_actual not in ("Auditor", "admin"):
        st.info("Solo Auditores")
    else:
        df_abiertos = listar_inventarios_abiertos()
        if df_abiertos.empty:
            st.info("No hay inventarios abiertos")
        else:
            id_sel = st.selectbox("Seleccionar inventario", df_abiertos["ID_Inventario"].astype(str).tolist())
            df_det = cargar_detalle(id_sel)
            if df_det.empty:
                st.warning("No hay detalle")
            else:
                cols_show = ["Concesionaria","Sucursal",C_LOC,C_ART,C_DESC,C_STOCK,C_COSTO,"Cat","Conteo_Fisico","Diferencia"]
                cols_show = [c for c in cols_show if c in df_det.columns]
                df_edit = df_det[cols_show].copy()

                edited = render_data_editor(
                    df_edit,
                    use_container_width=True,
                    num_rows="fixed",
                    disabled=[c for c in df_edit.columns if c != "Conteo_Fisico"],
                )

                df_det2 = df_det.copy()
                key_cols = [C_ART, C_LOC]
                if not all(c in df_det2.columns for c in key_cols):
                    st.error("Columnas no encontradas")
                else:
                    edited2 = edited.copy()
                    for c in key_cols:
                        edited2[c] = edited2[c].astype(str)
                        df_det2[c] = df_det2[c].astype(str)

                    df_merge = df_det2.merge(
                        edited2[key_cols + ["Conteo_Fisico"]],
                        on=key_cols,
                        how="left",
                        suffixes=("", "_new")
                    )

                    df_merge["Conteo_Fisico"] = df_merge["Conteo_Fisico_new"].combine_first(df_merge.get("Conteo_Fisico"))
                    if "Conteo_Fisico_new" in df_merge.columns:
                        df_merge = df_merge.drop(columns=["Conteo_Fisico_new"])

                    stock_num = parse_ar_number(df_merge[C_STOCK]).fillna(0)
                    conteo_num = parse_ar_number(df_merge["Conteo_Fisico"]).fillna(0)
                    df_merge["Diferencia"] = conteo_num - stock_num

                    col_save, col_dl = st.columns([1, 1])
                    with col_save:
                        if st.button("💾 Guardar conteo"):
                            ok = guardar_detalle_modificado(id_sel, df_merge)
                            if ok:
                                st.success("✅ Conteo guardado")
                            else:
                                st.error("Error al guardar conteo. Revisá Audit_Log o mensajes de error.")
                            st.rerun()

                    with col_dl:
                        df_view = df_edit.copy()
                        df_view["Conteo_Fisico"] = edited["Conteo_Fisico"]
                        stock_num_view = parse_ar_number(df_view[C_STOCK]).fillna(0)
                        conteo_num_view = parse_ar_number(df_view["Conteo_Fisico"]).fillna(0)
                        df_view["Diferencia"] = conteo_num_view - stock_num_view

                        cols_export = [c for c in cols_show if c in df_view.columns]
                        df_export = df_view[cols_export].copy()
                        xlsx_data = export_dataframe_to_excel(df_export, sheet_name="Conteo", title=f"Conteo Físico - {id_sel}")
                        st.download_button(
                            "⬇️ Descargar Conteo Excel",
                            data=xlsx_data,
                            file_name=f"Conteo_{id_sel}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

# ----------------------------
# MÓDULO 3
# ----------------------------
elif modulo_activo == "justificaciones":
    st.subheader("Justificaciones")

    df_abiertos = listar_inventarios_abiertos()
    if df_abiertos.empty:
        st.info("No hay inventarios abiertos")
    else:
        id_sel = st.selectbox("Seleccionar", df_abiertos["ID_Inventario"].astype(str).tolist(), key="tab3")
        df_det = cargar_detalle(id_sel)

        if df_det.empty:
            st.warning("No hay detalle")
        else:
            dif_num = pd.to_numeric(df_det.get("Diferencia", 0), errors="coerce").fillna(0)
            df_dif = df_det.loc[dif_num != 0].copy()

            if df_dif.empty:
                st.success("Sin diferencias")
            else:
                # Mostrar tabla resumen de conteos y diferencias
                st.write("### 📋 Resumen de Conteos y Diferencias")
                cols_resumen = [C_ART, C_LOC, C_STOCK, "Conteo_Fisico", "Diferencia", C_COSTO]
                cols_resumen = [c for c in cols_resumen if c in df_dif.columns]
                df_resumen = df_dif[cols_resumen].copy()
                render_dataframe(df_resumen, use_container_width=True, hide_index=True)
                st.divider()
                
                if rol_actual in ("Deposito", "admin"):
                    st.write("**Ingresá justificaciones:**")
                    justificaciones_dict = {}
                    
                    for idx, row in df_dif.iterrows():
                        art = row[C_ART]
                        loc = row[C_LOC]
                        dif = row["Diferencia"]
                        just_actual = row.get("Justificacion", "")
                        
                        st.write(f"**{art} ({loc}) - Diferencia: {dif}**")
                        just = st.text_area(
                            f"Justificación",
                            value=just_actual,
                            height=80,
                            key=f"just_{idx}"
                        )
                        justificaciones_dict[idx] = just
                        st.divider()
                    
                    if st.button("💾 Guardar justificaciones"):
                        df_det2 = df_det.copy()
                        for idx, just in justificaciones_dict.items():
                            df_det2.loc[df_det2.index == idx, "Justificacion"] = just
                        
                        ok = guardar_detalle_modificado(id_sel, df_det2)
                        if ok:
                            st.success("✅ Guardado")
                            # Opción de descargar
                            st.divider()
                            st.write("### 📥 Descargar justificaciones:")
                            cols_export = [C_ART, C_LOC, C_STOCK, "Conteo_Fisico", "Diferencia", "Justificacion", C_COSTO]
                            cols_export = [c for c in cols_export if c in df_det2.columns]
                            df_export = df_det2[cols_export].copy()
                            xlsx_data = export_dataframe_to_excel(df_export, sheet_name="Justificaciones", title=f"Justificaciones - {id_sel}")
                            st.download_button(
                                "⬇️ Descargar Justificaciones Excel",
                                data=xlsx_data,
                                file_name=f"Justificaciones_{id_sel}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.error("Error al guardar justificaciones. Revisá Audit_Log.")
                        st.rerun()
                else:
                    st.write("**Validá justificaciones y asignà ajustes:**")
                    validaciones_dict = {}
                    ajustes_dict = {}
                    canjes_invalidos = []
                    
                    for idx, row in df_dif.iterrows():
                        art = row[C_ART]
                        loc = row[C_LOC]
                        dif = row.get("Diferencia", 0)
                        costo = pd.to_numeric(row.get(C_COSTO, 0), errors="coerce")
                        just = row.get("Justificacion", "")
                        val_actual = row.get("Justif_Validada", "")
                        tipo_ajuste_actual = row.get("Tipo_Ajuste", "")
                        ajuste_cant_actual = row.get("Ajuste_Cantidad", "")
                        canje_codigo_actual = row.get("Canje_Articulo", "")
                        
                        st.write(f"**{art} ({loc}) - Diferencia: {dif} - Costo: {format_currency_ar(costo)}**")
                        st.write(f"*{just if just else '(sin justificación)'}*")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            val = st.selectbox(
                                "¿Validada?",
                                options=["", "SI", "NO"],
                                index=(["", "SI", "NO"].index(val_actual) if val_actual in ["SI", "NO"] else 0),
                                key=f"val_{idx}"
                            )
                        
                        # Habilitar ajuste solo si Justif_Validada es "SI"
                        if val == "SI":
                            with col2:
                                tipo_ajuste = st.selectbox(
                                    "Tipo de Ajuste",
                                    options=["", "Ajuste", "Canje", "Sin Ajuste"],
                                    index=(["", "Ajuste", "Canje", "Sin Ajuste"].index(tipo_ajuste_actual) if tipo_ajuste_actual in ["Ajuste", "Canje", "Sin Ajuste"] else 0),
                                    key=f"tipo_ajuste_{idx}"
                                )
                            
                            # Mostrar campo numérico solo si elige "Ajuste" o "Canje"
                            if tipo_ajuste in ("Ajuste", "Canje"):
                                ajuste_cant = st.number_input(
                                    f"Cantidad a {tipo_ajuste.lower()} (neg. faltante, pos. sobrante)",
                                    value=float(ajuste_cant_actual) if ajuste_cant_actual else 0.0,
                                    step=1.0,
                                    key=f"ajuste_cant_{idx}"
                                )
                            else:
                                ajuste_cant = 0.0
                                tipo_ajuste = "Sin Ajuste" if tipo_ajuste == "" else tipo_ajuste

                            canje_info = None
                            canje_codigo = ""
                            if tipo_ajuste == "Canje":
                                canje_codigo = st.text_input(
                                    "Código de artículo para canje",
                                    value=str(canje_codigo_actual) if canje_codigo_actual is not None else "",
                                    key=f"canje_codigo_{idx}",
                                    placeholder="Ingresá código de artículo",
                                )

                                canje_info = buscar_articulo_en_base(id_sel, canje_codigo)
                                if canje_codigo.strip() and canje_info:
                                    st.success("Artículo encontrado en la base del Excel importado")
                                    colc1, colc2, colc3, colc4 = st.columns(4)
                                    colc1.write(f"**Artículo:** {canje_info['codigo']}")
                                    colc2.write(f"**Descripción:** {canje_info['descripcion']}")
                                    colc3.write(f"**Costo Rep.:** {format_currency_ar(canje_info['costo'])}")
                                    colc4.write(f"**Stock base:** {canje_info['stock']:.2f}")
                                    st.write(f"**Cantidad a ajustar:** {ajuste_cant:.2f}")
                                elif canje_codigo.strip():
                                    st.error("Código no encontrado en la base completa del Excel importado.")
                                    canjes_invalidos.append(idx)
                        else:
                            tipo_ajuste = ""
                            ajuste_cant = 0.0
                            canje_codigo = ""
                            canje_info = None
                        
                        validaciones_dict[idx] = val
                        ajustes_dict[idx] = (tipo_ajuste, ajuste_cant, canje_codigo, canje_info)
                        st.divider()
                    
                    if st.button("💾 Guardar validación y ajustes"):
                        if canjes_invalidos:
                            st.error("Hay canjes con código inválido. Corregí los códigos antes de guardar.")
                            st.stop()

                        df_det2 = df_det.copy()
                        for idx, val in validaciones_dict.items():
                            df_det2.loc[df_det2.index == idx, "Justif_Validada"] = val
                            df_det2.loc[df_det2.index == idx, "Validador"] = usuario_actual
                            df_det2.loc[df_det2.index == idx, "Fecha_Validacion"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
                            if idx in ajustes_dict:
                                tipo, cantidad, canje_codigo, canje_info = ajustes_dict[idx]
                                df_det2.loc[df_det2.index == idx, "Tipo_Ajuste"] = tipo
                                df_det2.loc[df_det2.index == idx, "Ajuste_Cantidad"] = cantidad
                                if tipo == "Canje" and canje_info:
                                    df_det2.loc[df_det2.index == idx, "Canje_Articulo"] = canje_info["codigo"]
                                    df_det2.loc[df_det2.index == idx, "Canje_Descripcion"] = canje_info["descripcion"]
                                    df_det2.loc[df_det2.index == idx, "Canje_Costo_Rep"] = canje_info["costo"]
                                    df_det2.loc[df_det2.index == idx, "Canje_Stock_Base"] = canje_info["stock"]
                                else:
                                    df_det2.loc[df_det2.index == idx, "Canje_Articulo"] = ""
                                    df_det2.loc[df_det2.index == idx, "Canje_Descripcion"] = ""
                                    df_det2.loc[df_det2.index == idx, "Canje_Costo_Rep"] = ""
                                    df_det2.loc[df_det2.index == idx, "Canje_Stock_Base"] = ""
                        
                        ok = guardar_detalle_modificado(id_sel, df_det2)
                        if ok:
                            st.success("✅ Guardado")
                            # Opción de descargar
                            st.divider()
                            st.write("### 📥 Descargar validaciones y ajustes:")
                            cols_export = [C_ART, C_LOC, C_STOCK, "Conteo_Fisico", "Diferencia", "Justificacion", "Justif_Validada", "Tipo_Ajuste", "Ajuste_Cantidad", "Canje_Articulo", "Canje_Descripcion", "Canje_Costo_Rep", "Canje_Stock_Base", C_COSTO]
                            cols_export = [c for c in cols_export if c in df_det2.columns]
                            df_export = df_det2[cols_export].copy()
                            xlsx_data = export_dataframe_to_excel(df_export, sheet_name="Validaciones", title=f"Validaciones y Ajustes - {id_sel}")
                            st.download_button(
                                "⬇️ Descargar Validaciones Excel",
                                data=xlsx_data,
                                file_name=f"Validaciones_{id_sel}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.error("Error al guardar validaciones y ajustes. Revisá Audit_Log.")
                        st.rerun()

# ----------------------------
# MÓDULO 4
# ----------------------------
elif modulo_activo == "cierre":
    st.subheader("Cierre + Reporte")
    
    if rol_actual not in ("Auditor", "admin"):
        st.info("Solo Auditores")
    else:
        df_abiertos = listar_inventarios_abiertos()
        if df_abiertos.empty:
            st.info("No hay inventarios abiertos")
        else:
            id_sel = st.selectbox("Seleccionar para cerrar", df_abiertos["ID_Inventario"].astype(str).tolist(), key="tab4")
            df_det = cargar_detalle(id_sel)

            if df_det.empty:
                st.warning("No hay detalle")
            else:
                resultados = calcular_resultados_inventario(df_det)
                
                if not resultados:
                    st.error("Error en cálculos")
                    st.stop()
                
                st.write("### 📊 Resultados")
                
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Muestra", resultados["cant_muestra"])
                col2.metric("Faltantes", resultados["cant_faltantes"])
                col3.metric("Sobrantes", resultados["cant_sobrantes"])
                col4.metric("Grado", f"{resultados['grado']}%")
                
                # Calcular % de muestra para cada fila
                valor_muestra = resultados["valor_muestra"]
                tabla_resultados = pd.DataFrame([
                    {
                        "Detalle": "Muestra", 
                        "Q": resultados["cant_muestra"], 
                        "$ Ajuste": format_currency_ar(resultados['valor_muestra']),
                        "%": f"{resultados['pct_muestra']:.2f}%"
                    },
                    {
                        "Detalle": "Faltantes", 
                        "Q": resultados["cant_faltantes"], 
                        "$ Ajuste": format_currency_ar(resultados['valor_faltantes']),
                        "%": f"{resultados['pct_faltantes']:.2f}%"
                    },
                    {
                        "Detalle": "Sobrantes", 
                        "Q": resultados["cant_sobrantes"], 
                        "$ Ajuste": format_currency_ar(resultados['valor_sobrantes']),
                        "%": f"{resultados['pct_sobrantes']:.2f}%"
                    },
                    {
                        "Detalle": "Dif Neta", 
                        "Q": resultados["cant_dif_neta"], 
                        "$ Ajuste": format_currency_ar(resultados['valor_dif_neta']),
                        "%": f"{resultados['pct_dif_neta']:.2f}%"
                    },
                    {
                        "Detalle": "Dif Absoluta", 
                        "Q": resultados["cant_dif_absoluta"], 
                        "$ Ajuste": format_currency_ar(resultados['valor_dif_absoluta']),
                        "%": f"{resultados['pct_dif_absoluta']:.2f}%"
                    },
                ])
                
                render_dataframe(tabla_resultados, use_container_width=True, hide_index=True)
                
                # Mostrar resumen de canjes si existen
                if resultados.get("canjes"):
                    st.divider()
                    st.write("### 🔄 Resumen de Canjes")
                    df_canjes = pd.DataFrame(resultados["canjes"])
                    render_dataframe(df_canjes, use_container_width=True, hide_index=True)
                
                st.divider()
                st.write("### 📥 Descargar reporte:")

                def build_report_xlsx():
                    from openpyxl import Workbook
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

                    buffer = io.BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Resultado"

                    title_font = Font(size=14, bold=True)
                    light_red = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")
                    bold = Font(bold=True)
                    center = Alignment(horizontal="center", vertical="center")
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)

                    ws.merge_cells("A1:D1")
                    ws["A1"] = "4. Resultado Inventario Rotativo"
                    ws["A1"].font = title_font
                    ws["A1"].alignment = center

                    ws["A3"] = "Resultado:"

                    start_row = 5
                    ws[f"A{start_row}"] = "Detalle"
                    ws[f"B{start_row}"] = "Cant"
                    ws[f"C{start_row}"] = "$"
                    
                    rows = [
                        ("Muestra", resultados["cant_muestra"], resultados["valor_muestra"]),
                        ("Faltantes", resultados["cant_faltantes"], resultados["valor_faltantes"]),
                        ("Sobrantes", resultados["cant_sobrantes"], resultados["valor_sobrantes"]),
                        ("Dif Neta", resultados["cant_dif_neta"], resultados["valor_dif_neta"]),
                        ("Dif Absoluta", resultados["cant_dif_absoluta"], resultados["valor_dif_absoluta"]),
                    ]

                    for i, r in enumerate(rows, start=start_row + 1):
                        ws[f"A{i}"] = r[0]
                        ws[f"B{i}"] = r[1]
                        ws[f"C{i}"] = r[2]
                        # Apply Argentine format: . for thousands, , for decimals
                        ws[f"C{i}"].number_format = '#,##0.00'

                    ws2 = wb.create_sheet(title="Detalle")
                    # Detect numeric columns in detail sheet
                    numeric_cols = {}
                    for col_idx, col_name in enumerate(df_det.columns, start=1):
                        try:
                            numeric_test = pd.to_numeric(df_det[col_name], errors="coerce")
                            if numeric_test.notna().sum() > 0:
                                numeric_cols[col_idx] = col_name
                        except:
                            pass
                    
                    # Write detail sheet with formatting
                    for r_idx, row in enumerate(dataframe_to_rows(df_det, index=False, header=True), start=1):
                        for c_idx, value in enumerate(row, start=1):
                            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                            if r_idx == 1:
                                cell.font = Font(bold=True)
                            # Apply numeric formatting for numeric columns
                            if c_idx in numeric_cols:
                                try:
                                    num_val = float(value) if value not in (None, "", "NaN", "nan") else None
                                    if num_val is not None and str(num_val) not in ("nan", "inf", "-inf"):
                                        cell.value = num_val
                                        cell.number_format = '#,##0.00'
                                except:
                                    pass

                    wb.save(buffer)
                    buffer.seek(0)
                    return buffer

                xlsx_data = build_report_xlsx()
                st.download_button(
                    "⬇️ Descargar XLSX",
                    data=xlsx_data,
                    file_name=f"Reporte_{id_sel}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.divider()
                if st.button("✅ Cerrar inventario"):
                    cerrar_inventario(id_sel, usuario_actual)
                    st.success("Cerrado")
                    st.rerun()

# ----------------------------
# MÓDULO 5
# ----------------------------
elif modulo_activo == "dashboards":
    st.subheader("Dashboards")

    kpis = calcular_dashboard_kpis()

    fila1 = st.columns(4)
    fila1[0].metric("Inventarios Totales", kpis["inventarios_totales"])
    fila1[1].metric("Inventarios Abiertos", kpis["inventarios_abiertos"])
    fila1[2].metric("Inventarios Cerrados", kpis["inventarios_cerrados"])
    fila1[3].metric("Tasa de Cierre", f"{kpis['tasa_cierre']:.1f}%")

    fila2 = st.columns(3)
    fila2[0].metric("Líneas Muestreadas", kpis["lineas_muestreadas"])
    fila2[1].metric("Valuación Auditada", format_currency_ar(kpis['valuacion_muestra']))
    fila2[2].metric("Exactitud Promedio", f"{kpis['exactitud_promedio']:.1f}%")

    st.divider()
    st.write("### 📋 KPIs incluidos")
    render_dataframe(
        pd.DataFrame([
            {"KPI": "Inventarios Totales", "Descripción": "Cantidad total de inventarios generados en el sistema"},
            {"KPI": "Inventarios Abiertos", "Descripción": "Inventarios aún pendientes de cierre"},
            {"KPI": "Inventarios Cerrados", "Descripción": "Inventarios finalizados"},
            {"KPI": "Tasa de Cierre", "Descripción": "Porcentaje de inventarios cerrados sobre el total"},
            {"KPI": "Líneas Muestreadas", "Descripción": "Cantidad total de artículos incluidos en las muestras"},
            {"KPI": "Valuación Auditada", "Descripción": "Suma de Stock x Costo de Reposición de todas las muestras"},
            {"KPI": "Exactitud Promedio", "Descripción": "Promedio del grado de exactitud calculado en los inventarios"},
        ]),
        use_container_width=True,
        hide_index=True,
    )

    st.divider()
    st.write("### 🏢 Ranking de Sucursales")
    if not kpis["ranking_sucursales"].empty:
        render_dataframe(kpis["ranking_sucursales"], use_container_width=True, hide_index=True)
    else:
        st.info("Todavía no hay datos suficientes para mostrar ranking de sucursales.")

    st.divider()
    st.write("### 📦 Resumen por Inventario")
    if not kpis["detalle_resumen"].empty:
        render_dataframe(kpis["detalle_resumen"], use_container_width=True, hide_index=True)
    else:
        st.info("Todavía no hay inventarios con detalle para analizar.")
