import pandas as pd
import datetime
import io
from openpyxl import load_workbook

# Construir df_det de ejemplo con las columnas que usa la app
cols = [
    "ID_Inventario", "Concesionaria", "Sucursal", "Locación", "Artículo", "Descripción", "Stock", "Cto.Rep.", "Conteo_Fisico", "Diferencia"
]
rows = [
    ["INV-20260210-0001","Autolux","Ax Jujuy","H-01-09-A07","PH1860KB1000","CALCO LAT, SW4 - CAT 3",4,10000,1,-3],
    ["INV-20260210-0001","Autolux","Ax Jujuy","D-10-02-D03","851100K22200","MOTOR LIMPIAPARABRISAS",3,15000,1,-2],
    ["INV-20260210-0001","Autolux","Ax Jujuy","D-05-03-B06","480690D12100","PARRILLA SUSP,INF",1,50000,8,7],
    ["INV-20260210-0001","Autolux","Ax Jujuy","D-02-09-M01","V01-MIN-0001","JAULA INTERNA 4PTS",1,30000,2,1],
]

df_det = pd.DataFrame(rows, columns=cols)

# Copiar la función de report building desde app.py (ligeramente adaptada)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

C_STOCK = "Stock"
C_COSTO = "Cto.Rep."


def generate_report(df_det, id_sel, output_path):
    dfr = df_det.copy()
    stock_col = C_STOCK if C_STOCK in dfr.columns else None
    costo_col = C_COSTO if C_COSTO in dfr.columns else None
    dfr["_stock_num"] = pd.to_numeric(dfr.get(stock_col, 0), errors="coerce").fillna(0)
    dfr["_costo_num"] = pd.to_numeric(dfr.get(costo_col, 0), errors="coerce").fillna(0)
    dif = pd.to_numeric(dfr.get("Diferencia", 0), errors="coerce").fillna(0)

    muestra_cnt = int(dfr["_stock_num"].sum())
    valor_muestra = (dfr["_stock_num"] * dfr["_costo_num"]).sum()

    value_faltantes = ((-dif[dif < 0]) * dfr.loc[dif < 0, "_costo_num"]).sum()
    value_sobrantes = ((dif[dif > 0]) * dfr.loc[dif > 0, "_costo_num"]).sum()
    value_neta = (dif * dfr["_costo_num"]).sum()
    value_absoluta = (dif.abs() * dfr["_costo_num"]).sum()

    pct_absoluto = (abs(value_neta) / valor_muestra * 100) if valor_muestra else 0

    escala = [(0.00, 100), (0.10, 94), (0.80, 82), (1.60, 65), (2.40, 35), (3.30, 0)]
    escala_sorted = sorted(escala, key=lambda x: x[0])
    grado = 0
    for th, g in escala_sorted:
        if pct_absoluto >= th:
            grado = g

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    title_font = Font(size=14, bold=True)
    light_red = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    money_fmt = "#,##0.00"
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:D1")
    ws["A1"] = "4. Resultado Inventario Rotativo"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws["A3"] = "El resultado del inventario rotativo es el siguiente:"

    start_row = 5
    ws[f"A{start_row}"] = "Detalle"
    ws[f"B{start_row}"] = "Cant. de Art."
    ws[f"C{start_row}"] = "$"
    ws[f"D{start_row}"] = "%"
    for col in ["A","B","C","D"]:
        cell = ws[f"{col}{start_row}"]
        cell.font = bold
        cell.alignment = center
        cell.border = border

    rows = [
        ("Muestra", muestra_cnt, valor_muestra, 1.0),
        ("Faltantes", int(dfr.loc[dif < 0, "_stock_num"].sum()), value_faltantes, (value_faltantes / valor_muestra) if valor_muestra else 0),
        ("Sobrantes", int(dfr.loc[dif > 0, "_stock_num"].sum()), value_sobrantes, (value_sobrantes / valor_muestra) if valor_muestra else 0),
        ("Diferencia Neta", int(dfr["Diferencia"].sum()), value_neta, (value_neta / valor_muestra) if valor_muestra else 0),
        ("Diferencia Absoluta", int(dfr["Diferencia"].abs().sum()), value_absoluta, (value_absoluta / valor_muestra) if valor_muestra else 0),
    ]

    for i, r in enumerate(rows, start=start_row + 1):
        ws[f"A{i}"] = r[0]
        ws[f"B{i}"] = r[1]
        ws[f"C{i}"] = r[2]
        ws[f"D{i}"] = r[3]
        ws[f"B{i}"].alignment = center
        ws[f"C{i}"].number_format = money_fmt
        ws[f"D{i}"].number_format = "0.00%"
        ws[f"A{i}"].border = border
        ws[f"B{i}"].border = border
        ws[f"C{i}"].border = border
        ws[f"D{i}"].border = border

    diff_rows = [start_row + 3, start_row + 4]
    for r in diff_rows:
        for col in ["B","C","D"]:
            ws[f"{col}{r}"].fill = light_red

    pct_cell_row = start_row + 1
    ws.merge_cells(f"F{pct_cell_row}:G{pct_cell_row}")
    ws[f"F{pct_cell_row}"] = f"{grado}%"
    ws[f"F{pct_cell_row}"].font = Font(size=12, bold=True)
    ws[f"F{pct_cell_row}"].alignment = center

    escala_start = start_row + 7
    ws[f"B{escala_start}"] = "Dif. Abs. desde"
    ws[f"C{escala_start}"] = "Grado de cumplim."
    ws[f"B{escala_start}"].font = bold
    ws[f"C{escala_start}"].font = bold

    for j, (th, g) in enumerate(escala_sorted, start=escala_start + 1):
        ws[f"B{j}"] = f"{th:.2f}%"
        ws[f"C{j}"] = f"{g}%"
        ws[f"B{j}"].alignment = center
        ws[f"C{j}"].alignment = center

    ws2 = wb.create_sheet(title="Detalle")
    for r in dataframe_to_rows(df_det, index=False, header=True):
        ws2.append(r)

    for column_cells in ws.columns:
        try:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            col_letter = column_cells[0].column_letter
            ws.column_dimensions[col_letter].width = min(40, length + 4)
        except:
            pass

    wb.save(output_path)


if __name__ == '__main__':
    out = 'Reporte_test.xlsx'
    generate_report(df_det, 'INV-TEST', out)
    print('Reporte generado:', out)
